'''
ETL:    CDA2 Laboratory Report --> OMOP CDM v5.4
Usage:    python etl_cda2_omop.py
Output: Excel file with one sheet for each OMOP table
'''

import re
import os
import sys
from datetime import datetime
from lxml import etree
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# -----------------------------------------------------------------------------
# VOCAB LOOKUP - OMOPHub integration
# -----------------------------------------------------------------------------
from vocab_lookup import VocabLookup
_vocab = VocabLookup()

# Runtime cache
_organism_cache:   dict[str, int] = {}
_antibiotic_cache: dict[str, int] = {}
_specimen_cache:   dict[str, int] = {}
_culture_cache: dict[str, int] = {}


def lookup_organism(display_name: str) -> int:
    '''
    It receives the displayName of the organism from the CDA (e.g., Escherichia coli) 
    and returns the correct SNOMED concept_id via OMOPHub's semantic search.
    '''
    key = display_name.strip().lower()
    if key not in _organism_cache:
        _organism_cache[key] = _vocab.organism(display_name)
    return _organism_cache[key]


def lookup_antibiotic(display_name: str) -> int:
    '''
    It receives the display name of the antibiotic from the CDA (e.g., Amoxicillin/clavulanic acid) 
    and returns the LOINC concept_id for the corresponding MIC test via OMOPHub's semantic search.
    '''
    key = display_name.strip().lower()
    if key not in _antibiotic_cache:
        _antibiotic_cache[key] = _vocab.antibiotic(display_name)
    return _antibiotic_cache[key]


def lookup_specimen(display_name: str) -> int:
    '''
    It receives the displayName of the specimen type from the CDA 
    and returns the corresponding SNOMED concept_id. 
    Falls back to CONCEPT_SPECIMEN_UNK if 0.
    '''
    key = display_name.strip().lower()
    if key not in _specimen_cache:
        _specimen_cache[key] = _vocab.specimen(display_name)
    return _specimen_cache[key]


# -----------------------------------------------------------------------------
# STATIC MAPPING OF CULTURAL TESTS
# -----------------------------------------------------------------------------
# For this dataset, the types of exams are limited and known in advance.
def _norm_test_name(text: str) -> str:
    return re.sub(r'\s+', ' ', (text or '').strip().lower())


CULTURE_TEST_CONCEPT_MAP = {
    _norm_test_name('Esame colturale su broncolavaggio'): 1092251,
    _norm_test_name('Urinocoltura'): 3026008,
    _norm_test_name('EMOCOLTURA DA VENA PERIFERICA PER RICERCA GERMI AEROBI'): 3005745,
    _norm_test_name('EMOCOLTURA DA VENA PERIFERICA PER RICERCA GERMI ANAEROBI'): 3006673,
    _norm_test_name('Ricerca germi aerobi in bactec per controllo di sterilita terreno di conservazione'): 3013867,
    _norm_test_name('Ricerca germi anaerobi in bactec per controllo di sterilita terreno di conservazione'): 3011298,
    _norm_test_name('Esame colturale su espettorato'): 3023419,

    # Generic mappings.
    _norm_test_name('Valutazione attivita antibatterica farmaci in aerobiosi'): 3013867,
    _norm_test_name('Valutazione attivita antibatterica farmaci in anaerobiosi'): 3011298,
    _norm_test_name('Valutazione attivita antibatterica farmaci a temperatura ambiente'): 3013867,
    _norm_test_name('Ricerca germi aerobi per controllo di sterilita cellule staminali'): 3013867,
    _norm_test_name('Ricerca germi anaerobi per controllo di sterilita cellule staminali'): 3011298,
}


def lookup_culture_test(title: str, stype: str = '') -> int:
    '''
    Returns the static concept_id of the cultural test.
    '''
    key = _norm_test_name(title)
    concept_id = CULTURE_TEST_CONCEPT_MAP.get(key)
    if concept_id:
        print(f"  [CULTURE MAP] '{title}' -> {concept_id}")
        return concept_id

    if stype == 'aerobi':
        print(f"  [WARN CULTURE MAP] Title not listed: '{title}' -> fallback aerobe 3013867")
        return 3013867
    if stype == 'anaerobi':
        print(f"  [WARN CULTURE MAP] Title not listed: '{title}' -> fallback anaerobe 3011298")
        return 3011298

    print(f"  [WARN CULTURE MAP] Title not listed: '{title}' -> 0")
    return 0


# -----------------------------------------------------------------------------
# NAMESPACE CDA2
# -----------------------------------------------------------------------------
NS = {
    'hl7': 'urn:hl7-org:v3',
    'xsi': 'http://www.w3.org/2001/XMLSchema-instance'
}


# -----------------------------------------------------------------------------
# PHASE 1 - OMOP CDM 5.4 FIXED SCHEMA
# Define columns and order for each table.
# Fields marked as REQUIRED are mandatory: if they are missing, a warning is logged.
# -----------------------------------------------------------------------------
OMOP_SCHEMA = {
    'PERSON': [
        'person_id',                    # REQUIRED - PRIMARY KEY
        'gender_concept_id',            # REQUIRED - FOREIGN KEY
        'year_of_birth',                # REQUIRED
        'month_of_birth',
        'day_of_birth',
        'birth_datetime',
        'race_concept_id',              # REQUIRED - FOREIGN KEY (default 0)
        'ethnicity_concept_id',         # REQUIRED - FOREIGN KEY (default 0)
        'location_id',                  # FOREIGN KEY
        'provider_id',                  # FOREIGN KEY
        'care_site_id',                 # FOREIGN KEY
        'person_source_value',
        'gender_source_value',
        'gender_source_concept_id',     # FOREIGN KEY
        'race_source_value',
        'race_source_concept_id',       # FOREIGN KEY
        'ethnicity_source_value',
        'ethnicity_source_concept_id',
    ],
    'LOCATION': [
        'location_id',                  # REQUIRED - PRIMARY KEY
        'address_1',
        'address_2',
        'city',
        'state',
        'zip',
        'county',
        'location_source_value',
        'country_concept_id',           # FOREIGN KEY
        'country_source_value',
        'latitude',
        'longitude',
    ],
    'CARE_SITE': [
        'care_site_id',                 # REQUIRED - PRIMARY KEY
        'care_site_name',
        'place_of_service_concept_id',  # FOREIGN KEY
        'location_id',                  # FOREIGN KEY
        'care_site_source_value',
        'place_of_service_source_value',
    ],
    'VISIT_OCCURRENCE': [
        'visit_occurrence_id',          # REQUIRED - PRIMARY KEY
        'person_id',                    # REQUIRED - FOREIGN KEY
        'visit_concept_id',             # REQUIRED - FOREIGN KEY
        'visit_start_date',             # REQUIRED
        'visit_start_datetime',
        'visit_end_date',               # REQUIRED
        'visit_end_datetime',
        'visit_type_concept_id',        # REQUIRED - FOREIGN KEY
        'provider_id',                  # FOREIGN KEY
        'care_site_id',                 # FOREIGN KEY
        'visit_source_value',
        'visit_source_concept_id',      # FOREIGN KEY
        'admitted_from_concept_id',     # FOREIGN KEY
        'admitted_from_source_value',
        'discharged_to_concept_id',     # FOREIGN KEY
        'discharged_to_source_value',
        'preceding_visit_occurrence_id',# FOREIGN KEY
    ],
    'SPECIMEN': [
        'specimen_id',                  # REQUIRED - PRIMARY KEY
        'person_id',                    # REQUIRED - FOREIGN KEY
        'specimen_concept_id',          # REQUIRED - FOREIGN KEY
        'specimen_type_concept_id',     # REQUIRED - FOREIGN KEY
        'specimen_date',                # REQUIRED
        'specimen_datetime',
        'quantity',
        'unit_concept_id',              # FOREIGN KEY
        'anatomic_site_concept_id',     # FOREIGN KEY
        'disease_status_concept_id',    # FOREIGN KEY
        'specimen_source_id',
        'specimen_source_value',
        'unit_source_value',
        'anatomic_site_source_value',
        'disease_status_source_value',
    ],
    'MEASUREMENT': [
        'measurement_id',               # REQUIRED - PRIMARY KEY
        'person_id',                    # REQUIRED - FOREIGN KEY
        'measurement_concept_id',       # REQUIRED - FOREIGN KEY
        'measurement_date',             # REQUIRED
        'measurement_datetime',
        'measurement_time',
        'measurement_type_concept_id',  # REQUIRED - FOREIGN KEY
        'operator_concept_id',          # FOREIGN KEY
        'value_as_number',
        'value_as_concept_id',          # FOREIGN KEY
        'unit_concept_id',              # FOREIGN KEY
        'range_low',
        'range_high',
        'provider_id',                  # FOREIGN KEY
        'visit_occurrence_id',          # FOREIGN KEY
        'visit_detail_id',              # FOREIGN KEY
        'measurement_source_value',
        'measurement_source_concept_id',# FOREIGN KEY
        'unit_source_value',
        'unit_source_concept_id',       # FOREIGN KEY
        'value_source_value',
        'measurement_event_id',
        'meas_event_field_concept_id',  # FOREIGN KEY
    ],
    'OBSERVATION': [
        'observation_id',               # REQUIRED - PRIMARY KEY
        'person_id',                    # REQUIRED - FOREIGN KEY
        'observation_concept_id',       # REQUIRED - FOREIGN KEY
        'observation_date',             # REQUIRED
        'observation_datetime',
        'observation_type_concept_id',  # REQUIRED - FOREIGN KEY
        'value_as_number',
        'value_as_string',
        'value_as_concept_id',          # FOREIGN KEY
        'qualifier_concept_id',         # FOREIGN KEY
        'unit_concept_id',              # FOREIGN KEY
        'provider_id',                  # FOREIGN KEY
        'visit_occurrence_id',          # FOREIGN KEY
        'visit_detail_id',              # FOREIGN KEY
        'observation_source_value',
        'observation_source_concept_id',# FOREIGN KEY
        'unit_source_value',
        'qualifier_source_value',
        'value_source_value',
        'observation_event_id',
        'obs_event_field_concept_id',   # FOREIGN KEY
    ],
}

# Required fields for each table (used for validation)
REQUIRED_FIELDS = {
    'PERSON':           ['person_id', 'gender_concept_id', 'year_of_birth',
                         'race_concept_id', 'ethnicity_concept_id'],
    'LOCATION':         ['location_id'],
    'CARE_SITE':        ['care_site_id'],
    'VISIT_OCCURRENCE': ['visit_occurrence_id', 'person_id', 'visit_concept_id',
                         'visit_start_date', 'visit_end_date', 'visit_type_concept_id'],
    'SPECIMEN':         ['specimen_id', 'person_id', 'specimen_concept_id',
                         'specimen_type_concept_id', 'specimen_date'],
    'MEASUREMENT':      ['measurement_id', 'person_id', 'measurement_concept_id',
                         'measurement_date', 'measurement_type_concept_id'],
    'OBSERVATION':      ['observation_id', 'person_id', 'observation_concept_id',
                         'observation_date', 'observation_type_concept_id'],
}


# -----------------------------------------------------------------------------
# LOOKUP TABLES
# -----------------------------------------------------------------------------
GENDER_MAP = {
    'M': 8507,
    'F': 8532,
}

RESULT_MAP = {
    'Positivo': 9191,
    'Negativo': 9189,
    'POS':      9191,
    'NEG':      9189,
}

INTERP_MAP = {
    'R': 4148441,   # Resistant
    'S': 4038110,   # Susceptible
    'I': 4043352,   # Intermediate
}

ANTIBIOGRAM_STRING_MAP = {
    'neg':   9189,   # Negative
    'pos':   9191,   # Positive
}

SPECIMEN_CODE_MAP = {
    'SAE':  618898,   # Blood specimen in blood culture bottle
    'URMI': 4122280,  # Mid-stream urine specimen
    'SAN':  618898,   # Sangue (fallback generico)
}


# -----------------------------------------------------------------------------
# CONCEPT ID OMOP (fixed)
# -----------------------------------------------------------------------------
CONCEPT_LAB              = 32856
CONCEPT_LAB_VISIT        = 32036
CONCEPT_DETECTED         = 4126681
CONCEPT_MEAS_MID         = 1147729
CONCEPT_OBS_OID          = 1147762
CONCEPT_SPEC_ID          = 1147049
CONCEPT_GRAM_STAIN       = 3037167
CONCEPT_AEROBE_CULTURE   = 3013867
CONCEPT_ANAEROBE_CULTURE = 3011298
CONCEPT_SPECIMEN_UNK     = 4002873
CONCEPT_ITALY            = 41987173
UNIT_MGL                 = 8751
UNIT_HOUR                = 8505
UNIT_MINUTE              = 8550


# -----------------------------------------------------------------------------
# ID GENERATOR
# Each table has its own dedicated sequence with a specific starting value.
# In production, replace with the database logic.
# -----------------------------------------------------------------------------
ID_START = {
    'LOCATION':         99,
    'CARE_SITE':        99999,
    'VISIT_OCCURRENCE': 199999,
    'PERSON':           399999,
    'SPECIMEN':         499999,
    'MEASUREMENT':      699999,
    'OBSERVATION':      899999,
}

_id_counters = ID_START.copy()


def new_id(entity: str) -> int:
    global _id_counters
    if entity not in _id_counters:
        raise ValueError(f"Entità ID non gestita: '{entity}'")
    _id_counters[entity] += 1
    return _id_counters[entity]


def reset_id_counter():
    global _id_counters
    _id_counters = ID_START.copy()


# -----------------------------------------------------------------------------
# PHASE 2 - HELPER: make_record
# Creates a record with all schema columns initialized to None.
# Accepts only keys present in the schema; otherwise, raises a ValueError.
# Logs a warning if a REQUIRED field is None after assignment.
# -----------------------------------------------------------------------------
def make_record(table: str, **values) -> dict:
    record = {col: None for col in OMOP_SCHEMA[table]}
    for key, val in values.items():
        if key not in record:
            raise ValueError(f"[SCHEMA] Colonna '{key}' non esiste in {table}")
        record[key] = val
    # Validazione campi obbligatori
    for req in REQUIRED_FIELDS.get(table, []):
        if record.get(req) is None:
            print(f"  [WARN] {table}.{req} is REQUIRED, but the value is None")
    return record


# -----------------------------------------------------------------------------
# HELPER: preprocessing CDA
# -----------------------------------------------------------------------------
def preprocess_cda(content: str) -> str:
    # Fix 1: scientific notation in attributes
    def fix_attr(m):
        full = m.group(0)
        attr = m.group(1)
        val  = m.group(2)
        tz   = ''
        base = val
        tz_m = re.search(r'([+-]\d{4})$', val)
        if tz_m:
            tz   = tz_m.group(1)
            base = val[:tz_m.start()]
        try:
            if re.search(r'[eE]', base):
                fixed = str(int(float(base))) + tz
                return f'{attr}="{fixed}"'
        except Exception:
            pass
        return full
    content = re.sub(r'([\w:]+)="([^"]*[eE][^"]*)"', fix_attr, content)

    # Fix 2: < used as an operator (e.g., <=0.25) within the text of an element.
    content = re.sub(r'(?<=>)(\s*)(<=|>=|<|>)(\s*\d)', 
                     lambda m: m.group(1) + m.group(2).replace('<', '&lt;').replace('>', '&gt;') + m.group(3), 
                     content)

    # Fix 3: unwanted whitespace in the closing tag: < /tag> --> </tag>
    content = re.sub(r'<\s+/', '</', content)

    return content


# -----------------------------------------------------------------------------
# HELPER: datetime CDA2 conversion
# -----------------------------------------------------------------------------
def parse_cda_datetime(raw: str):
    if not raw:
        return None, None
    clean = re.sub(r'[+-]\d{4}$', '', raw)
    try:
        if len(clean) >= 14:
            dt = datetime(int(clean[0:4]), int(clean[4:6]), int(clean[6:8]),
                          int(clean[8:10]), int(clean[10:12]), int(clean[12:14]))
            return dt.date(), dt
        elif len(clean) >= 8:
            dt = datetime(int(clean[0:4]), int(clean[4:6]), int(clean[6:8]))
            return dt.date(), dt
    except Exception:
        pass
    return None, None


# -----------------------------------------------------------------------------
# HELPER: TTP ("0d 0h 2min") --> hour float
# -----------------------------------------------------------------------------
def parse_ttp_to_hours(s: str):
    if not s:
        return None
    days  = int(m.group(1)) if (m := re.search(r'(\d+)d',   s)) else 0
    hours = int(m.group(1)) if (m := re.search(r'(\d+)h',   s)) else 0
    mins  = int(m.group(1)) if (m := re.search(r'(\d+)min', s)) else 0
    total = days * 24 + hours + mins / 60
    return round(total, 4) if total > 0 else None


# -----------------------------------------------------------------------------
# HELPER: TTP ("0d 0h 2min") --> minutes (my choice)
# -----------------------------------------------------------------------------
def parse_ttp_to_minute(s:str):
    if not s:
        return None
    days  = int(m.group(1)) if (m := re.search(r'(\d+)d',   s)) else 0
    hours = int(m.group(1)) if (m := re.search(r'(\d+)h',   s)) else 0
    mins  = int(m.group(1)) if (m := re.search(r'(\d+)min', s)) else 0
    total = days * 1440 + hours * 60 + mins
    return total if total > 0 else None


# -----------------------------------------------------------------------------
# HELPER: XPath shortcut
# -----------------------------------------------------------------------------
def xp(node, path: str, default=None):
    result = node.xpath(path, namespaces=NS)
    return result[0] if result else default


# -----------------------------------------------------------------------------
# SECTION TYPE
# -----------------------------------------------------------------------------
def classify_section(code: str, title: str) -> str:
    code_l = code.lower()
    title_l = title.lower()

    if re.search(r'attivit[aà]\s*antibatteric', title_l) or re.search(r'antsb', code_l):
        return 'attivita_antibatterica'
    if re.search(r'curoc|urin', code_l) or 'urinocoltura' in title_l:
        return 'urinocoltura'
    if re.search(r'boana|ceman|cemn', code_l) or 'anaerobi' in title_l:
        return 'anaerobi'
    if re.search(r'boaer|cemae|ceme', code_l) or 'aerobi' in title_l:
        return 'aerobi'
    if re.search(r'^tp\d', code_l) or 'positivizz' in title_l or 'tempo di posit' in title_l:
        return 'ttp'
    if re.search(r'esmic', code_l) or 'microscopico' in title_l or 'esame microscopico' in title_l:
        return 'microscopia'
    if 'esame colturale' in title_l:
        return 'coltura'
    return 'unknown'


# -----------------------------------------------------------------------------
# Set count
# -----------------------------------------------------------------------------
def get_set_key(title: str, code: str) -> str:
    '''
    Retrieves the key for the set from the section. 
    Returns ‘set1’, ‘set2’, etc. Works with both a single set and multiple sets.
    '''
    title_l = title.lower()
    # Explicit patterns
    if re.search(r'\bii\s*set\b|set\s*ii\b|2[°º]\s*set|set\s*2\b', title_l):
        return 'set2'
    if re.search(r'\biii\s*set\b|set\s*iii\b|3[°º]\s*set|set\s*3\b', title_l):
        return 'set3'
    # Default: If there is no multiple-set marker, it is set 1
    return 'set1'

# -----------------------------------------------------------------------------
# ORGANISM IDENTIFICATION + ANTIBIOGRAM
# -----------------------------------------------------------------------------
def process_organisms_and_antibiogram(omop, section, code, meas_date, meas_dt,
                                       person_id, visit_id, mid, current_spec_id):

    current_organism_oid = None

    for cluster in section.xpath('.//hl7:organizer[@classCode="CLUSTER"]', namespaces=NS):
        org_code = xp(cluster, 'hl7:specimen/hl7:specimenRole/hl7:specimenPlayingEntity/hl7:code/@code', '')
        org_name = xp(cluster, 'hl7:specimen/hl7:specimenRole/hl7:specimenPlayingEntity/hl7:code/@displayName', '')

        if not org_code or org_code.upper() in ('USUB',):
            continue

        oid = new_id('OBSERVATION')
        current_organism_oid = oid

        org_concept_id = lookup_organism(org_name.strip() if org_name else org_code)

        omop['OBSERVATION'].append(make_record('OBSERVATION',
            observation_id=oid,
            person_id=person_id,
            observation_concept_id=org_concept_id,
            observation_date=meas_date,
            observation_datetime=meas_dt,
            observation_type_concept_id=CONCEPT_LAB,
            value_as_string=org_name or None,
            value_as_concept_id=CONCEPT_DETECTED,
            visit_occurrence_id=visit_id,
            observation_source_value=org_code,
            observation_source_concept_id=0,
            value_source_value=org_name or org_code or None,
            observation_event_id=mid,
            obs_event_field_concept_id=CONCEPT_MEAS_MID,
        ))

        # ── Antibiogram (organizer BATTERY inside CLUSTER) ──────────
        for battery in cluster.xpath('.//hl7:organizer[@classCode="BATTERY"]', namespaces=NS):
            for comp in battery.xpath('hl7:component/hl7:observation', namespaces=NS):
                ab_code = xp(comp, 'hl7:code/@code', '')
                ab_name = xp(comp, 'hl7:code/@displayName', '')
                ab_interp = xp(comp, 'hl7:interpretationCode/@code', '')

                if not ab_code or ab_code.upper() == code.upper():
                    continue

                ab_concept_id = lookup_antibiotic(ab_name.strip() if ab_name else ab_code)

                ab_val_raw = xp(comp, 'hl7:value/@value', '')
                ab_val_text = xp(comp, 'hl7:value/text()', '') or ''
                ab_val_text = re.sub(r'\s+', '', ab_val_text)
                ab_val_text = ab_val_text.replace('≤', '<=').replace('≥', '>=')
                ab_unit = xp(comp, 'hl7:value/@unit', '')

                operator_concept_id = None
                value_as_number = None
                value_as_concept_id = INTERP_MAP.get(ab_interp)
                unit_concept_id = None
                unit_source_value = None

                # CASE 1: Neg - Pos
                if not ab_val_raw and ab_val_text.lower() in ANTIBIOGRAM_STRING_MAP:
                    value_as_concept_id = ANTIBIOGRAM_STRING_MAP[ab_val_text.lower()]
                    omop['MEASUREMENT'].append(make_record('MEASUREMENT',
                        measurement_id=new_id('MEASUREMENT'),
                        person_id=person_id,
                        measurement_concept_id=ab_concept_id,
                        measurement_date=meas_date,
                        measurement_datetime=meas_dt,
                        measurement_type_concept_id=CONCEPT_LAB,
                        value_as_concept_id=value_as_concept_id,
                        visit_occurrence_id=visit_id,
                        measurement_source_value=ab_code,
                        measurement_source_concept_id=0,
                        value_source_value=ab_val_text,
                        measurement_event_id=oid,
                        meas_event_field_concept_id=CONCEPT_OBS_OID,
                    ))
                    continue

                # CASE 2: value string (no numbers)
                if (not ab_val_raw and ab_val_text
                        and not re.match(r'^\s*(<=|>=|<|>|=)?\s*[\d]+(?:[.,][\d]+)?\s*$', ab_val_text)):
                    omop['MEASUREMENT'].append(make_record('MEASUREMENT',
                        measurement_id=new_id('MEASUREMENT'),
                        person_id=person_id,
                        measurement_concept_id=ab_concept_id,
                        measurement_date=meas_date,
                        measurement_datetime=meas_dt,
                        measurement_type_concept_id=CONCEPT_LAB,
                        value_as_concept_id=INTERP_MAP.get(ab_interp),
                        visit_occurrence_id=visit_id,
                        measurement_source_value=ab_code,
                        measurement_source_concept_id=0,
                        value_source_value=ab_val_text,
                        measurement_event_id=oid,
                        meas_event_field_concept_id=CONCEPT_OBS_OID,
                    ))
                    continue

                # CASO 3: Numeric value (with or without operator)
                if ab_val_raw or ab_val_text:
                    if ab_val_raw:
                        try:
                            value_as_number = float(ab_val_raw.replace(',', '.'))
                        except ValueError:
                            value_as_number = None
                        operator_concept_id = 4172703  # =
                    elif ab_val_text:
                        ab_val_text_norm = re.sub(r'\s+', '', ab_val_text)
                        ab_val_text_norm = ab_val_text_norm.replace('≤', '<=').replace('≥', '>=')
                        m = re.match(r'^(<=|>=|<|>|=)?([\d]+(?:[.,][\d]+)?)$', ab_val_text_norm)
                        if m:
                            op_str = m.group(1) or '='
                            try:
                                value_as_number = float(m.group(2).replace(',', '.'))
                            except ValueError:
                                value_as_number = None
                            operator_concept_id = {
                                '<=': 4171754, '<': 4171754,
                                '>=': 4172704, '>': 4172704,
                                '=': 4172703,
                            }.get(op_str, None)

                    unit_concept_id = UNIT_MGL if value_as_number is not None else None
                    unit_source_value = ab_unit or ('mg/L' if value_as_number is not None else None)

                    omop['MEASUREMENT'].append(make_record('MEASUREMENT',
                        measurement_id=new_id('MEASUREMENT'),
                        person_id=person_id,
                        measurement_concept_id=ab_concept_id,
                        measurement_date=meas_date,
                        measurement_datetime=meas_dt,
                        measurement_type_concept_id=CONCEPT_LAB,
                        operator_concept_id=operator_concept_id,
                        value_as_number=value_as_number,
                        value_as_concept_id=INTERP_MAP.get(ab_interp),
                        unit_concept_id=unit_concept_id,
                        visit_occurrence_id=visit_id,
                        measurement_source_value=ab_code,
                        measurement_source_concept_id=0,
                        unit_source_value=unit_source_value,
                        value_source_value=ab_val_raw or ab_val_text or None,
                        measurement_event_id=oid,
                        meas_event_field_concept_id=CONCEPT_OBS_OID,
                    ))

    return current_organism_oid

# -----------------------------------------------------------------------------
# PRINCIPAL PARSER
# -----------------------------------------------------------------------------
def parse_cda2_to_omop(cda_path: str) -> dict:

    with open(cda_path, 'r', encoding='utf-8') as f:
        content = f.read()
    content = preprocess_cda(content)
    root    = etree.fromstring(content.encode('utf-8'))

    omop = {table: [] for table in OMOP_SCHEMA}

    # 1. PERSON
    person_id    = new_id('PERSON')
    gender_code  = xp(root, '//hl7:patient/hl7:administrativeGenderCode/@code', '')
    birth_raw    = xp(root, '//hl7:patient/hl7:birthTime/@value', '')
    cf_val       = xp(root, '//hl7:patientRole/hl7:id[@assigningAuthorityName="Ministero Economia e Finanze"]/@extension', '')
    lab_id       = xp(root, '//hl7:patientRole/hl7:id[@assigningAuthorityName="Laboratorio Analisi"]/@extension', '')

    birth_date, birth_datetime = parse_cda_datetime(birth_raw)

    person_record = make_record('PERSON',
        person_id            = person_id,
        gender_concept_id    = GENDER_MAP.get(gender_code, 0),
        gender_source_value  = gender_code or None,
        year_of_birth        = int(birth_raw[0:4]) if len(birth_raw) >= 4 else None,
        month_of_birth       = int(birth_raw[4:6]) if len(birth_raw) >= 6 else None,
        day_of_birth         = int(birth_raw[6:8]) if len(birth_raw) >= 8 else None,
        birth_datetime       = birth_datetime,
        race_concept_id      = 0,
        ethnicity_concept_id = 0,
        person_source_value  = lab_id or cf_val or None,
        # location_id and care_site_id: added later (if any)
    )
    omop['PERSON'].append(person_record)

    # 2. LOCATION (patient)
    pat_streets = root.xpath('//hl7:patientRole/hl7:addr/hl7:streetAddressLine/text()', namespaces=NS)
    pat_cities  = root.xpath('//hl7:patientRole/hl7:addr/hl7:city/text()', namespaces=NS)
    pat_zips    = root.xpath('//hl7:patientRole/hl7:addr/hl7:postalCode/text()', namespaces=NS)

    if any([pat_streets, pat_cities, pat_zips]):
        loc_id_pat = new_id('LOCATION')
        omop['LOCATION'].append(make_record('LOCATION',
            location_id          = loc_id_pat,
            address_1            = pat_streets[0] if pat_streets else None,
            city                 = pat_cities[0]  if pat_cities  else None,
            zip                  = pat_zips[0]    if pat_zips    else None,
            country_concept_id   = CONCEPT_ITALY,
            country_source_value = 'ITA',
        ))
        person_record['location_id'] = loc_id_pat

    # 3. LOCATION (care site) + CARE_SITE
    org_names   = root.xpath('//hl7:serviceEvent//hl7:representedOrganization/hl7:name/text()', namespaces=NS)
    org_streets = root.xpath('//hl7:serviceEvent//hl7:representedOrganization/hl7:addr/hl7:streetAddressLine/text()', namespaces=NS)
    org_cities  = root.xpath('//hl7:serviceEvent//hl7:representedOrganization/hl7:addr/hl7:city/text()', namespaces=NS)
    org_zips    = root.xpath('//hl7:serviceEvent//hl7:representedOrganization/hl7:addr/hl7:postalCode/text()', namespaces=NS)

    if org_names or any([org_streets, org_cities, org_zips]):
        loc_id_cs   = new_id('LOCATION')
        care_sit_id = new_id('CARE_SITE')

        omop['LOCATION'].append(make_record('LOCATION',
            location_id          = loc_id_cs,
            address_1            = org_streets[0] if org_streets else None,
            city                 = org_cities[0]  if org_cities  else None,
            zip                  = org_zips[0]    if org_zips    else None,
            country_concept_id   = CONCEPT_ITALY,
            country_source_value = 'ITA',
        ))
        omop['CARE_SITE'].append(make_record('CARE_SITE',
            care_site_id                  = care_sit_id,
            care_site_name                = org_names[0] if org_names else None,
            place_of_service_concept_id   = 0,
            location_id                   = loc_id_cs,
            place_of_service_source_value = org_names[0] if org_names else None,
        ))
        person_record['care_site_id'] = care_sit_id
    else:
        care_sit_id = None

    # 4. VISIT_OCCURRENCE ───────────────────────────────────────────────────
    visit_id    = new_id('VISIT_OCCURRENCE')
    v_start_raw = xp(root, '//hl7:encompassingEncounter/hl7:effectiveTime/hl7:low/@value',  '')
    v_end_raw   = xp(root, '//hl7:encompassingEncounter/hl7:effectiveTime/hl7:high/@value', '')
    order_val   = xp(root, '//hl7:order/hl7:id/@extension', '')

    v_start_d, v_start_dt = parse_cda_datetime(v_start_raw)
    v_end_d,   v_end_dt   = parse_cda_datetime(v_end_raw)

    omop['VISIT_OCCURRENCE'].append(make_record('VISIT_OCCURRENCE',
        visit_occurrence_id     = visit_id,
        person_id               = person_id,
        visit_concept_id        = CONCEPT_LAB_VISIT,
        visit_start_date        = v_start_d,
        visit_start_datetime    = v_start_dt,
        visit_end_date          = v_end_d,
        visit_end_datetime      = v_end_dt,
        visit_type_concept_id   = CONCEPT_LAB,
        care_site_id            = care_sit_id,
        visit_source_value      = order_val or None,
        visit_source_concept_id = 0,
    ))

    # SPECIMEN
    specimen_map: dict[str, tuple] = {}

    SPECIMEN_CODE_MAP = {
        'SAE':  618898,   # Blood specimen in blood culture bottle
        'URMI': 4122280,  # Mid-stream urine specimen
        'SAN':  618898,   # Sangue (fallback generico)
    }

    batt_sections = root.xpath('//hl7:section[hl7:code/@code="18719-5"]/hl7:component/hl7:section', namespaces=NS)

    def _specimen_key_for_section(sec_code: str, sec_title: str, sec_type: str) -> str:
        return f"set|{get_set_key(sec_title, sec_code)}"

    for _sec in batt_sections:
        _code  = xp(_sec, 'hl7:code/@code', '')
        _title = xp(_sec, 'hl7:title/text()', '').strip()
        _stype = classify_section(_code, _title)

        if _stype not in ('aerobi', 'anaerobi', 'coltura', 'attivita_antibatterica'):
            continue

        _spec_key = _specimen_key_for_section(_code, _title, _stype)

        if _spec_key not in specimen_map:
            _spec_date_raw = xp(_sec, 'hl7:entry/hl7:act/hl7:effectiveTime/@value', '') or \
                             xp(_sec, './/hl7:observation/hl7:effectiveTime/@value', v_start_raw)
            _spec_d, _spec_dt = parse_cda_datetime(_spec_date_raw)

            _spec_code_raw = xp(
                _sec,
                'hl7:entry/hl7:act/hl7:specimen/hl7:specimenRole/'
                'hl7:specimenPlayingEntity/hl7:code/@code',
                ''
            )

            _translation_code = xp(
                _sec,
                'hl7:entry/hl7:act/hl7:specimen/hl7:specimenRole/'
                'hl7:specimenPlayingEntity/hl7:code/hl7:translation/@code',
                ''
            )
            _translation_display = xp(
                _sec,
                'hl7:entry/hl7:act/hl7:specimen/hl7:specimenRole/'
                'hl7:specimenPlayingEntity/hl7:code/hl7:translation/@displayName',
                ''
            )

            if _spec_code_raw.upper() == 'USUB':
                _spec_concept = CONCEPT_SPECIMEN_UNK
            else:
                _spec_concept = SPECIMEN_CODE_MAP.get(_translation_code) or \
                                (lookup_specimen(_translation_display) if _translation_display else 0) or \
                                CONCEPT_SPECIMEN_UNK

            _sid = new_id('SPECIMEN')
            _spec_lookup_display = _translation_display or _spec_code_raw or _translation_code or ''
            specimen_map[_spec_key] = (_sid, _spec_lookup_display)

            omop['SPECIMEN'].append(make_record('SPECIMEN',
                specimen_id               = _sid,
                person_id                 = person_id,
                specimen_concept_id       = _spec_concept,
                specimen_type_concept_id  = CONCEPT_LAB,
                specimen_date             = _spec_d,
                specimen_datetime         = _spec_dt,
                specimen_source_id        = _translation_code or None,
                specimen_source_value     = _translation_display or _spec_code_raw or _translation_code or _spec_key,
                anatomic_site_concept_id  = 0,
                disease_status_concept_id = 0,
            ))

    # 6. LOOP
    current_culture_mid  = None
    current_spec_id      = None
    current_organism_oid = None
    meas_date = v_start_d
    meas_dt   = v_start_dt

    for section in batt_sections:
        code  = xp(section, 'hl7:code/@code', '')
        title = xp(section, 'hl7:title/text()', '').strip()
        stype = classify_section(code, title)

        # Culture
        if stype in ('aerobi', 'anaerobi', 'coltura'):
            mid = new_id('MEASUREMENT')
            current_culture_mid = mid

            current_spec_key = _specimen_key_for_section(code, title, stype)
            _spec_tuple = specimen_map.get(current_spec_key)
            current_spec_id = _spec_tuple[0] if _spec_tuple else None
            _spec_display_curr = _spec_tuple[1] if _spec_tuple else ''

            meas_concept = lookup_culture_test(title, stype)

            obs_nodes = section.xpath(
                './/hl7:observation[hl7:code/@code=$codeval]/hl7:value/text()',
                namespaces=NS, codeval=code
            )
            val_text = obs_nodes[0] if obs_nodes else ''

            omop['MEASUREMENT'].append(make_record('MEASUREMENT',
                measurement_id=mid,
                person_id=person_id,
                measurement_concept_id=meas_concept,
                measurement_date=meas_date,
                measurement_datetime=meas_dt,
                measurement_type_concept_id=CONCEPT_LAB,
                value_as_concept_id=RESULT_MAP.get(val_text),
                visit_occurrence_id=visit_id,
                measurement_source_value=code,
                measurement_source_concept_id=0,
                value_source_value=val_text or None,
                measurement_event_id=current_spec_id,
                meas_event_field_concept_id=CONCEPT_SPEC_ID if current_spec_id else None,
            ))

            current_organism_oid = process_organisms_and_antibiogram(
                omop, section, code, meas_date, meas_dt, person_id, visit_id, mid, current_spec_id
            )

        elif stype == 'attivita_antibatterica':
            aerobic = 'aerobiosi' in title.lower() or 'temperatura ambiente' in title.lower()
            culture_type = 'aerobe' if aerobic else 'anaerobe'

            current_spec_key = _specimen_key_for_section(code, title, stype)
            _spec_tuple = specimen_map.get(current_spec_key)
            current_spec_id = _spec_tuple[0] if _spec_tuple else None

            meas_concept = lookup_culture_test(title, 'aerobi' if aerobic else 'anaerobi')

            mid = new_id('MEASUREMENT')
            current_culture_mid = mid

            obs_nodes = section.xpath(
                './/hl7:observation[hl7:code/@code=$codeval]/hl7:value/text()',
                namespaces=NS, codeval=code
            )
            val_text = obs_nodes[0] if obs_nodes else ''

            omop['MEASUREMENT'].append(make_record('MEASUREMENT',
                measurement_id=mid,
                person_id=person_id,
                measurement_concept_id=meas_concept,
                measurement_date=meas_date,
                measurement_datetime=meas_dt,
                measurement_type_concept_id=CONCEPT_LAB,
                value_as_concept_id=RESULT_MAP.get(val_text),
                visit_occurrence_id=visit_id,
                measurement_source_value=code,
                measurement_source_concept_id=0,
                value_source_value=val_text or None,
                measurement_event_id=current_spec_id,
                meas_event_field_concept_id=CONCEPT_SPEC_ID if current_spec_id else None,
            ))

            current_organism_oid = process_organisms_and_antibiogram(
                omop, section, code, meas_date, meas_dt, person_id, visit_id, mid, current_spec_id
            )

        # ── TTP ───────────────────────────────────
        elif stype == 'ttp':
            ttp_raw = xp(section, './/hl7:value/text()', '')
            ttp_m   = parse_ttp_to_minute(ttp_raw)

            omop['MEASUREMENT'].append(make_record('MEASUREMENT',
                measurement_id                = new_id('MEASUREMENT'),
                person_id                     = person_id,
                measurement_concept_id        = 3332492,   # Test Duration
                measurement_date              = meas_date,
                measurement_datetime          = meas_dt,
                measurement_type_concept_id   = CONCEPT_LAB,
                value_as_number               = ttp_m,
                unit_concept_id               = UNIT_MINUTE,
                visit_occurrence_id           = visit_id,
                measurement_source_value      = code,
                measurement_source_concept_id = 0,
                unit_source_value             = None,
                value_source_value            = ttp_raw or None,
                measurement_event_id          = current_culture_mid,
                meas_event_field_concept_id   = CONCEPT_MEAS_MID if current_culture_mid else None,
            ))

        # Microscopy exam
        elif stype == 'microscopia':
            mic_val = xp(section, './/hl7:value/text()', '')

            omop['OBSERVATION'].append(make_record('OBSERVATION',
                observation_id                = new_id('OBSERVATION'),
                person_id                     = person_id,
                observation_concept_id        = CONCEPT_GRAM_STAIN,
                observation_date              = meas_date,
                observation_datetime          = meas_dt,
                observation_type_concept_id   = CONCEPT_LAB,
                value_as_string               = mic_val or None,
                visit_occurrence_id           = visit_id,
                observation_source_value      = code,
                observation_source_concept_id = 0,
                value_source_value            = mic_val or None,
                observation_event_id          = current_culture_mid,
                obs_event_field_concept_id    = CONCEPT_MEAS_MID if current_culture_mid else None,
            ))

        # Urine colture
        elif stype == 'urinocoltura':
            uro_val = xp(section, './/hl7:value/text()', '')
            uro_val = uro_val.replace("\n", "")

            uro_spec_date_raw = xp(section, 'hl7:entry/hl7:act/hl7:effectiveTime/@value', v_start_raw)
            uro_spec_d, uro_spec_dt = parse_cda_datetime(uro_spec_date_raw)

            uro_spec_code_raw = xp(section, './/hl7:specimenPlayingEntity/hl7:code/@code', '')
            uro_mat_code = xp(section,
                './/hl7:specimenPlayingEntity/hl7:code/hl7:translation[@codeSystemName="MATERIALE"]/@code', '')
            uro_spec_display = xp(section,
                './/hl7:specimenPlayingEntity/hl7:code/hl7:translation[@codeSystemName="MATERIALE"]/@displayName', '')

            uro_specimen_code_map = {
                'SAE':  618898,   # Blood specimen in blood culture bottle
                'URMI': 4122280,  # Mid-stream urine specimen
                'SAN':  618898,
            }

            uro_spec_concept = uro_specimen_code_map.get(uro_mat_code) or \
                               (lookup_specimen(uro_spec_display) if uro_spec_display else 0) or \
                               CONCEPT_SPECIMEN_UNK

            uro_sid = new_id('SPECIMEN')
            omop['SPECIMEN'].append(make_record('SPECIMEN',
                specimen_id               = uro_sid,
                person_id                 = person_id,
                specimen_concept_id       = uro_spec_concept,
                specimen_type_concept_id  = CONCEPT_LAB,
                specimen_date             = uro_spec_d,
                specimen_datetime         = uro_spec_dt,
                specimen_source_id        = f'{uro_spec_code_raw}/{uro_mat_code}' if uro_mat_code else uro_spec_code_raw or None,
                specimen_source_value     = uro_spec_display or uro_mat_code or 'urinocoltura',
                anatomic_site_concept_id  = 0,
                disease_status_concept_id = 0,
            ))

            uro_meas_concept = lookup_culture_test(title or 'Urinocoltura', 'urinocoltura')
            omop['MEASUREMENT'].append(make_record('MEASUREMENT',
                measurement_id                = new_id('MEASUREMENT'),
                person_id                     = person_id,
                measurement_concept_id        = uro_meas_concept,
                measurement_date              = meas_date,
                measurement_datetime          = meas_dt,
                measurement_type_concept_id   = CONCEPT_LAB,
                value_as_concept_id           = RESULT_MAP.get((uro_val or '').strip()),
                visit_occurrence_id           = visit_id,
                measurement_source_value      = code,
                measurement_source_concept_id = 0,
                value_source_value            = re.sub(' +', ' ', uro_val) if uro_val else None,
                measurement_event_id          = uro_sid,
                meas_event_field_concept_id   = CONCEPT_SPEC_ID,
            ))

        else:
            print(f'    [WARN] Section not classified: code="{code}", title="{title}"')

    return omop


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 3 — EXCEL
# ─────────────────────────────────────────────────────────────────────────────
def write_excel(omop: dict, output_path: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font  = Font(name='Calibri', bold=True, color='FFFFFF')
    header_fill  = PatternFill(fill_type='solid', fgColor='01696F')
    header_align = Alignment(horizontal='center', vertical='center')

    for table, records in omop.items():
        ws      = wb.create_sheet(title=table)
        headers = OMOP_SCHEMA[table]

        for col_idx, header in enumerate(headers, start=1):
            cell           = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align

        for row_idx, record in enumerate(records, start=2):
            for col_idx, key in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=record.get(key))

        for col in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        print(f"  Sheet '{table}' --> {len(records)} record(s)")

    wb.save(output_path)
    print(f"\n  Saved file: {output_path}")

# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("=" * 60)
    print("  ETL: CDA2 Laboratory Report --> OMOP CDM 5.4")
    print("=" * 60)

    cda_file = input("\nInserisci il percorso del file CDA2 (es. C:/referti/referto.xml):\n> ").strip()
    cda_file = os.path.normpath(cda_file)

    if not os.path.isfile(cda_file):
        print(f"\n[ERRORE] File not found: {cda_file}")
        sys.exit(1)

    default_output = os.path.join(os.path.dirname(cda_file), 'output_omop.xlsx')
    output_path = input(f"\nOutput Excel file path (Default value: {default_output}):\n> ").strip()
    if not output_path:
        output_path = default_output
    output_path = os.path.normpath(output_path)

    print(f"\n[ETL] Parsing: {cda_file}")
    reset_id_counter()

    try:
        omop_data = parse_cda2_to_omop(cda_file)
    except Exception as e:
        print(f"\n[ERROR] Parsing failed: {e}")
        sys.exit(1)

    print("\n[ETL] Records:")
    for table, records in omop_data.items():
        print(f"  {table:<25} {len(records)} record(s)")

    print(f"\n[ETL] Writing Excel: {output_path}")
    write_excel(omop_data, output_path)
    print("\n[ETL] Done.")

    input("\nPress Enter to exit...")